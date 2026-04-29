#!/usr/bin/env python3
"""Convert the 10mila Körschema Excel file into schedule.csv.

Cell background colors in column C are pivoted into a `category` text
column so the webpage can present them without needing the original
Excel styling.
"""

from __future__ import annotations

import csv
import glob
import sys
from pathlib import Path

import openpyxl


COLOR_CATEGORIES: dict[str, str] = {
    "FFFFFF00": "Skylt",
    "FFFFC000": "Förvarning",
    "FFFF0000": "Mål",
    "FF00B0F0": "Växling",
    "FF0070C0": "Jaktstart",
    "THEME3_LIGHT": "Admin",
}

CATEGORY_DEFAULT = "Tidplan"


def cell_color_key(cell) -> str | None:
    fill = cell.fill
    if not fill or not fill.fgColor:
        return None
    fg = fill.fgColor
    if fg.type == "rgb":
        rgb = fg.rgb or ""
        if rgb in ("00000000", "FFFFFFFF", ""):
            return None
        return rgb.upper()
    if fg.type == "theme":
        if fg.theme == 3 and (fg.tint or 0) > 0.5:
            return "THEME3_LIGHT"
    return None


def category_for(cell) -> str:
    key = cell_color_key(cell)
    if key and key in COLOR_CATEGORIES:
        return COLOR_CATEGORIES[key]
    return CATEGORY_DEFAULT


def fmt_time(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat(timespec="minutes")
    return str(value)


def find_xlsx() -> Path:
    here = Path(__file__).resolve().parent
    matches = sorted(here.glob("*.xlsx"))
    if not matches:
        sys.exit("No .xlsx file found next to convert.py")
    return matches[0]


def main() -> None:
    xlsx_path = find_xlsx()
    out_path = xlsx_path.parent / "schedule.csv"

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    rows = []
    header_row = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        c = ws.cell(row=r, column=3).value
        if isinstance(a, str) and a.strip().lower().startswith("tidpunkt"):
            header_row = r
            break
    if header_row is None:
        sys.exit("Could not locate header row (Tidpunkt) in the sheet.")

    for r in range(header_row + 1, ws.max_row + 1):
        time_cell = ws.cell(row=r, column=1)
        adj_cell = ws.cell(row=r, column=2)
        task_cell = ws.cell(row=r, column=3)
        notes_cell = ws.cell(row=r, column=4)
        responsible_cell = ws.cell(row=r, column=7)
        location_cell = ws.cell(row=r, column=9)

        time_val = time_cell.value
        task_val = task_cell.value
        if time_val is None and not task_val:
            continue
        if task_val is None:
            continue

        rows.append(
            {
                "time": fmt_time(time_val),
                "adjusted_time": fmt_time(adj_cell.value),
                "category": category_for(task_cell),
                "task": str(task_val).strip(),
                "notes": ("" if notes_cell.value is None else str(notes_cell.value).strip()),
                "responsible": (
                    "" if responsible_cell.value is None else str(responsible_cell.value).strip()
                ),
                "location": (
                    "" if location_cell.value is None else str(location_cell.value).strip()
                ),
            }
        )

    rows.sort(key=lambda x: (x["adjusted_time"] or x["time"]))

    fieldnames = [
        "time",
        "adjusted_time",
        "category",
        "task",
        "notes",
        "responsible",
        "location",
    ]
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    print(f"Wrote {len(rows)} rows to {out_path}")


if __name__ == "__main__":
    main()
